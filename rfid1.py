import serial
import os
from openpyxl import load_workbook
from time import sleep
from time import gmtime, strftime


x = os.getcwd()
filepath="%s/demo.xlsx" %x
wb=load_workbook(filepath)
sheet_add=wb['Sheet1']

def add_excel_name():
	sheet_add['B2']="Hasnat"
	sheet_add['C2']=strftime("%Y-%m-%d %H:%M:%S", gmtime())
	
	r=2
	e = sheet_add.cell(row=r,column=2)
	j = str(e.value)
	while not j=="None":
		r+=1
		e = sheet_add.cell(row=r,column=2)
		j = str(e.value)
	e.value = "Hasib"
	
	wb.save(filepath)
	
def add_excel_time():
	r=2
	e = sheet_add.cell(row=r,column=3)
	j = str(e.value)
	while not j=="None":
		r+=1
		e = sheet_add.cell(row=r,column=3)
		j = str(e.value)
	e.value = strftime("%Y-%m-%d %H:%M:%S", gmtime())
	

COM = '/dev/tty.usbmodem1411'
BAUD = 9600

ser = serial.Serial(COM, BAUD, timeout = .1)

print('Waiting for device');
sleep(2)
print(ser.name)


while True:
	val = str(ser.readline().decode().strip('\r\n'))#Capture serial output as a decoded string
	#print()
	
	if (val == "24850484852515349555654563"):
		print("Hasnat", end="\r", flush=True)
		add_excel_name()
		add_excel_time()
